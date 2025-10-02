#!/usr/bin/env python3
"""
Parse fio output files (JSON or text) from a directory and export a readable Excel summary.

Usage examples:
  - Parse current directory and write results.xlsx
      python parse_fio.py

  - Specify input directory and output path
      python parse_fio.py --input /path/to/fio/results --output fio_summary.xlsx

  - Restrict to certain extensions and recurse
      python parse_fio.py --include json,txt,log --recurse

It extracts per-job metrics from JSON outputs (preferred) and best-effort metrics from text logs:
  - jobname, rw, bs, iodepth, numjobs, runtime
  - For read/write: iops, bandwidth, latency mean, clat p99 (when available)
"""

import argparse
import json
import math
import os
import re
import sys
from dataclasses import dataclass
from typing import Dict, Iterable, List, Optional, Tuple
from datetime import datetime

try:
    # For adding charts to Excel sheets
    from openpyxl.chart import BarChart, Reference, LineChart
    from openpyxl.formatting.rule import ColorScaleRule
except Exception:
    BarChart = None
    Reference = None
    LineChart = None
    ColorScaleRule = None

try:
    import pandas as pd
except ImportError as exc:
    sys.stderr.write(
        "Missing dependency: pandas. Install requirements with: pip install -r requirements.txt\n"
    )
    raise


# -----------------------
# Helpers and data models
# -----------------------

@dataclass
class ParsedRow:
    source_file: str
    task: Optional[str]
    timestamp: Optional[datetime]
    runner: Optional[str]
    jobname: Optional[str]
    rw: Optional[str]
    access_pattern: Optional[str]
    bs: Optional[str]
    bs_bytes: Optional[int]
    iodepth: Optional[int]
    numjobs: Optional[int]
    runtime_s: Optional[float]
    size_bytes: Optional[int]
    op: str  # 'read' or 'write'
    iops: Optional[float]
    bw_MBps: Optional[float]
    lat_mean_ms: Optional[float]
    clat_p99_ms: Optional[float]
    clat_p95_ms: Optional[float]
    clat_p50_ms: Optional[float]
    total_ios: Optional[int]


def _safe_div(num: float, den: float) -> Optional[float]:
    try:
        if den == 0:
            return None
        return num / den
    except Exception:
        return None


def _to_float(value: Optional[float]) -> Optional[float]:
    if value is None:
        return None
    try:
        return float(value)
    except Exception:
        return None


def _ns_to_ms(ns_value: Optional[float]) -> Optional[float]:
    if ns_value is None:
        return None
    return ns_value / 1_000_000.0


def _us_to_ms(us_value: Optional[float]) -> Optional[float]:
    if us_value is None:
        return None
    return us_value / 1_000.0


def _kbps_to_MBps(kb_per_s: Optional[float]) -> Optional[float]:
    if kb_per_s is None:
        return None
    # fio reports bw in KiB/s (1024). Convert to MB/s (10^6) for user familiarity.
    # Many prefer MiB/s; if you do, change divisor to 1024.
    return (kb_per_s * 1024.0) / 1_000_000.0


def _parse_size_to_bytes(value: Optional[str]) -> Optional[int]:
    if not value:
        return None
    try:
        # Accept inputs like "4G", "128M", "32k", or raw integer bytes as string
        match = re.match(r"^(\d+)([KkMmGgTtPp]?)[Bb]?$", value.strip())
        if not match:
            return int(value)
        num = int(match.group(1))
        unit = match.group(2).lower()
        power = {"": 0, "k": 10, "m": 20, "g": 30, "t": 40, "p": 50}.get(unit, 0)
        return num * (1 << power)
    except Exception:
        return None


def _parse_bs_to_bytes(value: Optional[str]) -> Optional[int]:
    # Alias to size parser; BS strings look similar (e.g., 4k, 256k)
    return _parse_size_to_bytes(value)


def _normalize_access_pattern(rw: Optional[str]) -> Optional[str]:
    if not rw:
        return None
    rw_l = str(rw).lower()
    if rw_l.startswith("rand"):
        return "random"
    if rw_l in ("read", "write", "rw", "readwrite"):
        return "sequential"
    return None


# -----------------------
# JSON parser
# -----------------------

def parse_fio_json(data: Dict, source_file: str, task: Optional[str], runner: Optional[str]) -> List[ParsedRow]:
    jobs = data.get("jobs") or []
    rows: List[ParsedRow] = []
    for job in jobs:
        jobname = job.get("jobname")
        options = job.get("job options") or {}
        rw = options.get("rw") or options.get("readwrite")
        bs = options.get("bs")
        iodepth = _to_int(options.get("iodepth"))
        numjobs = _to_int(options.get("numjobs"))
        # Preferred runtime from job, fallback to global
        runtime_ms = job.get("runtime") or data.get("global options", {}).get("runtime")
        runtime_s = _to_float(runtime_ms) if runtime_ms is not None else None
        if runtime_s is not None and runtime_s > 1000:
            # Some outputs store milliseconds
            runtime_s = runtime_s / 1000.0

        # Size
        size_bytes = _to_int(job.get("job options", {}).get("size"))
        if size_bytes is None:
            size_bytes = _parse_size_to_bytes(job.get("job options", {}).get("size"))

        for op in ("read", "write"):
            if op not in job:
                continue
            op_data = job[op]
            iops = _to_float(op_data.get("iops"))
            bw_kib_s = _to_float(op_data.get("bw"))  # in KiB/s
            bw_MBps = _kbps_to_MBps(bw_kib_s)
            total_ios = _to_int(op_data.get("total_ios"))

            # Latency: detect unit containers explicitly
            clat_ns = op_data.get("clat_ns")
            clat_us = op_data.get("clat_us")
            clat_generic = op_data.get("clat")
            lat_ns = op_data.get("lat_ns")
            lat_us = op_data.get("lat_us")
            lat_generic = op_data.get("lat")

            lat_mean_ms = None
            if isinstance(clat_ns, dict) and "mean" in clat_ns:
                lat_mean_ms = _ns_to_ms(_to_float(clat_ns.get("mean")))
            elif isinstance(clat_us, dict) and "mean" in clat_us:
                lat_mean_ms = _us_to_ms(_to_float(clat_us.get("mean")))
            elif isinstance(clat_generic, dict) and "mean" in clat_generic:
                # fio v2 JSON often uses microseconds for 'clat'
                lat_mean_ms = _us_to_ms(_to_float(clat_generic.get("mean")))
            elif isinstance(lat_ns, dict) and "mean" in lat_ns:
                lat_mean_ms = _ns_to_ms(_to_float(lat_ns.get("mean")))
            elif isinstance(lat_us, dict) and "mean" in lat_us:
                lat_mean_ms = _us_to_ms(_to_float(lat_us.get("mean")))
            elif isinstance(lat_generic, dict) and "mean" in lat_generic:
                lat_mean_ms = _us_to_ms(_to_float(lat_generic.get("mean")))

            # P99
            clat_p99_ms = None
            clat_p95_ms = None
            clat_p50_ms = None
            def _extract_pcts(container: Optional[Dict], to_ms_fn):
                nonlocal clat_p99_ms, clat_p95_ms, clat_p50_ms
                if not isinstance(container, dict):
                    return
                percentiles = container.get("percentile") or container.get("percentiles")
                if isinstance(percentiles, dict):
                    p99_key = _find_percentile_key(percentiles, 99.0)
                    if p99_key is not None:
                        clat_p99_ms = to_ms_fn(_to_float(percentiles.get(p99_key)))
                    p95_key = _find_percentile_key(percentiles, 95.0)
                    if p95_key is not None:
                        clat_p95_ms = to_ms_fn(_to_float(percentiles.get(p95_key)))
                    p50_key = _find_percentile_key(percentiles, 50.0)
                    if p50_key is not None:
                        clat_p50_ms = to_ms_fn(_to_float(percentiles.get(p50_key)))

            if clat_ns is not None:
                _extract_pcts(clat_ns, _ns_to_ms)
            elif clat_us is not None:
                _extract_pcts(clat_us, _us_to_ms)
            elif clat_generic is not None:
                _extract_pcts(clat_generic, _us_to_ms)

            rows.append(
                ParsedRow(
                    source_file=os.path.basename(source_file),
                    task=task,
                    timestamp=compute_timestamp_from_path(source_file),
                    runner=runner,
                    jobname=jobname,
                    rw=rw,
                    access_pattern=_normalize_access_pattern(rw),
                    bs=bs,
                    bs_bytes=_parse_bs_to_bytes(bs),
                    iodepth=iodepth,
                    numjobs=numjobs,
                    runtime_s=runtime_s,
                    size_bytes=size_bytes,
                    op=op,
                    iops=iops,
                    bw_MBps=bw_MBps,
                    lat_mean_ms=lat_mean_ms,
                    clat_p99_ms=clat_p99_ms,
                    clat_p95_ms=clat_p95_ms,
                    clat_p50_ms=clat_p50_ms,
                    total_ios=total_ios,
                )
            )
    return rows


def _to_int(value: Optional[object]) -> Optional[int]:
    if value is None:
        return None
    try:
        return int(value)
    except Exception:
        # handle strings like "4k", "1m"
        try:
            return int(float(str(value)))
        except Exception:
            return None


def _find_percentile_key(percentiles: Dict[str, object], target: float) -> Optional[str]:
    # fio uses keys like "99.000000"
    if str(target) in percentiles:
        return str(target)
    formatted = f"{target:.6f}"
    if formatted in percentiles:
        return formatted
    # try to find numerically closest key
    try:
        as_pairs = []
        for key in percentiles.keys():
            try:
                as_pairs.append((abs(float(key) - target), key))
            except Exception:
                continue
        if not as_pairs:
            return None
        as_pairs.sort()
        return as_pairs[0][1]
    except Exception:
        return None


# -----------------------
# Text parser (best-effort)
# -----------------------

_RE_IOPS_BW = re.compile(
    r"(?P<op>read|write):.*?IOPS\s*=\s*(?P<iops>[\d\.]+)\s*,\s*BW\s*=\s*(?P<bw_value>[\d\.]+)\s*(?P<bw_unit>(?:[KMG]i?)?B/s)",
    re.IGNORECASE,
)

_RE_LAT_AVG = re.compile(
    r"\b(?:clat|lat)\s*\([^)]*\):.*?avg\s*=\s*(?P<avg>[\d\.]+)\s*(?P<unit>n?sec|usec|msec|ms|us|ns)",
    re.IGNORECASE | re.DOTALL,
)

_RE_META = re.compile(
    r"\brw=(?P<rw>\w+).*?bs=(?P<bs>\S+)\b.*?(?:iodepth|iodepth_batch)=(?P<iodepth>\d+)",
    re.IGNORECASE,
)


def _parse_bw_to_MBps(value: str, unit: str) -> Optional[float]:
    unit = unit.strip()
    try:
        v = float(value)
    except Exception:
        return None
    unit_norm = unit.lower()
    # Handle KB/s, KiB/s, MB/s, MiB/s, GB/s, GiB/s
    if unit_norm in ("kb/s", "kib/s"):
        return (v * 1024) / 1_000_000.0
    if unit_norm in ("mb/s", "mib/s"):
        return (v * 1024 * 1024) / 1_000_000.0
    if unit_norm in ("gb/s", "gib/s"):
        return (v * 1024 * 1024 * 1024) / 1_000_000.0
    if unit_norm.endswith("b/s"):
        # bytes per second
        return v / 1_000_000.0
    return None


def _parse_latency_to_ms(value: str, unit: str) -> Optional[float]:
    try:
        v = float(value)
    except Exception:
        return None
    u = unit.lower()
    if u in ("ns", "nsec"):
        return v / 1_000_000.0
    if u in ("us", "usec"):
        return v / 1_000.0
    if u in ("ms", "msec"):
        return v
    return None


def parse_fio_text(text: str, source_file: str, task: Optional[str], runner: Optional[str]) -> List[ParsedRow]:
    rows: List[ParsedRow] = []

    # Meta
    rw = None
    bs = None
    iodepth = None
    meta_match = _RE_META.search(text)
    if meta_match:
        rw = meta_match.group("rw")
        bs = meta_match.group("bs")
        iodepth = _to_int(meta_match.group("iodepth"))

    # Latency (use the first avg found as a generic latency)
    lat_mean_ms_global: Optional[float] = None
    lat_match = _RE_LAT_AVG.search(text)
    if lat_match:
        lat_mean_ms_global = _parse_latency_to_ms(lat_match.group("avg"), lat_match.group("unit"))

    # IOPS/BW per op
    for m in _RE_IOPS_BW.finditer(text):
        op = m.group("op").lower()
        iops = _to_float(m.group("iops"))
        bw_MBps = _parse_bw_to_MBps(m.group("bw_value"), m.group("bw_unit"))
        rows.append(
            ParsedRow(
                source_file=os.path.basename(source_file),
                task=task,
                timestamp=compute_timestamp_from_path(source_file),
                runner=runner,
                jobname=None,
                rw=rw,
                access_pattern=_normalize_access_pattern(rw),
                bs=bs,
                bs_bytes=_parse_bs_to_bytes(bs),
                iodepth=iodepth,
                numjobs=None,
                runtime_s=None,
                size_bytes=None,
                op=op,
                iops=iops,
                bw_MBps=bw_MBps,
                lat_mean_ms=lat_mean_ms_global,
                clat_p99_ms=None,
                clat_p95_ms=None,
                clat_p50_ms=None,
            )
        )

    # If we didn't find explicit read/write lines, try to create a single generic row
    if not rows and (lat_mean_ms_global is not None):
        rows.append(
            ParsedRow(
                source_file=os.path.basename(source_file),
                task=task,
                timestamp=compute_timestamp_from_path(source_file),
                runner=runner,
                jobname=None,
                rw=rw,
                access_pattern=_normalize_access_pattern(rw),
                bs=bs,
                bs_bytes=_parse_bs_to_bytes(bs),
                iodepth=iodepth,
                numjobs=None,
                runtime_s=None,
                size_bytes=None,
                op="unknown",
                iops=None,
                bw_MBps=None,
                lat_mean_ms=lat_mean_ms_global,
                clat_p99_ms=None,
                clat_p95_ms=None,
                clat_p50_ms=None,
            )
        )

    return rows


# -----------------------
# Discovery and export
# -----------------------

DEFAULT_EXTENSIONS = ("json", "txt", "log", "out")


def iter_files(root: str, include_exts: Iterable[str], recurse: bool) -> Iterable[str]:
    include_lower = {ext.lower().lstrip(".") for ext in include_exts}
    if recurse:
        for dirpath, _dirnames, filenames in os.walk(root):
            for fname in filenames:
                ext = os.path.splitext(fname)[1].lower().lstrip(".")
                if ext in include_lower:
                    yield os.path.join(dirpath, fname)
    else:
        for fname in os.listdir(root):
            fpath = os.path.join(root, fname)
            if not os.path.isfile(fpath):
                continue
            ext = os.path.splitext(fname)[1].lower().lstrip(".")
            if ext in include_lower:
                yield fpath


_RE_TASK = re.compile(r"^parallel-\d+-[a-z]+-", re.IGNORECASE)


def compute_task_name(path: str, input_root: str) -> Optional[str]:
    try:
        rel = os.path.relpath(path, input_root)
    except Exception:
        rel = path
    parts = rel.split(os.sep)
    for part in parts:
        if _RE_TASK.match(part):
            return part
    # Fallback: first directory component if available
    if len(parts) > 1:
        return parts[0]
    return None


_RE_TIMESTAMP = re.compile(r"(20\d{6}T\d{6}Z)")


def compute_timestamp_from_path(path: str) -> Optional[datetime]:
    try:
        m = _RE_TIMESTAMP.search(path)
        if not m:
            # Try basename only
            m = _RE_TIMESTAMP.search(os.path.basename(path))
        if not m:
            return None
        ts = m.group(1)
        # Format: YYYYMMDDTHHMMSSZ
        return datetime.strptime(ts, "%Y%m%dT%H%M%SZ")
    except Exception:
        return None


_RE_RUNNER = re.compile(r"^fio-runner-[a-z0-9]+$", re.IGNORECASE)


def compute_runner_name(path: str, input_root: str) -> Optional[str]:
    try:
        rel = os.path.relpath(path, input_root)
    except Exception:
        rel = path
    for part in rel.split(os.sep):
        if _RE_RUNNER.match(part):
            return part
    return None


def sanitize_sheet_name(name: str, used: Optional[Dict[str, int]] = None) -> str:
    # Excel constraints: <=31 chars, no : \ / ? * [ ]
    invalid = set(':\\/?*[]')
    safe = ''.join(c if c not in invalid else '_' for c in name)
    safe = safe[:31]
    if used is None:
        return safe
    base = safe
    idx = used.get(base, 0)
    if idx == 0 and base not in used:
        used[base] = 1
        return base
    # Ensure uniqueness by appending a numeric suffix
    while True:
        idx += 1
        candidate = base[: max(0, 31 - len(f"_{idx}"))] + f"_{idx}"
        if candidate not in used:
            used[candidate] = 1
            return candidate


_RE_PARALLEL_DEGREE = re.compile(r"parallel-(\d+)", re.IGNORECASE)


def _extract_parallel_degree(task_name: Optional[str]) -> Optional[int]:
    if not task_name or not isinstance(task_name, str):
        return None
    m = _RE_PARALLEL_DEGREE.search(task_name)
    if not m:
        return None
    try:
        return int(m.group(1))
    except Exception:
        return None


def _sort_df_by_task_order(df: "pd.DataFrame", task_col: str = "task") -> "pd.DataFrame":
    if task_col not in df.columns:
        return df
    desired = [3, 6, 9, 12, 15]
    def parts(name: Optional[str]):
        deg = _extract_parallel_degree(name)
        if deg in desired:
            return (desired.index(deg), deg)
        # place non-listed/unknown after the known ones, ordered by degree then name
        return (len(desired), deg if deg is not None else 10**9)

    # Reset index to avoid pandas PerformanceWarning on MultiIndex operations
    tmp = df.reset_index(drop=True).copy()
    tmp["__ord1__"], tmp["__ord2__"] = zip(*tmp[task_col].map(parts))
    tmp = tmp.sort_values(by=["__ord1__", "__ord2__", task_col], kind="mergesort")
    # Remove helper columns without triggering MultiIndex drop warnings
    if "__ord1__" in tmp.columns:
        tmp.pop("__ord1__")
    if "__ord2__" in tmp.columns:
        tmp.pop("__ord2__")
    return tmp


def parse_file(path: str, input_root: str) -> List[ParsedRow]:
    try:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            content = f.read()
    except Exception as exc:
        sys.stderr.write(f"Failed to read {path}: {exc}\n")
        return []

    # Try JSON first
    try:
        data = json.loads(content)
        task = compute_task_name(path, input_root)
        runner = compute_runner_name(path, input_root)
        return parse_fio_json(data, source_file=path, task=task, runner=runner)
    except Exception:
        pass

    # Fallback to text parser
    try:
        task = compute_task_name(path, input_root)
        runner = compute_runner_name(path, input_root)
        return parse_fio_text(content, source_file=path, task=task, runner=runner)
    except Exception as exc:
        sys.stderr.write(f"Failed to parse text {path}: {exc}\n")
        return []


def rows_to_dataframe(rows: List[ParsedRow]) -> "pd.DataFrame":
    if not rows:
        return pd.DataFrame()
    records = [r.__dict__ for r in rows]
    df = pd.DataFrame.from_records(records)
    preferred_order = [
        "source_file",
        "task",
        "timestamp",
        "runner",
        "jobname",
        "op",
        "rw",
        "access_pattern",
        "bs",
        "bs_bytes",
        "iodepth",
        "numjobs",
        "runtime_s",
        "size_bytes",
        "iops",
        "bw_MBps",
        "lat_mean_ms",
        "clat_p99_ms",
        "clat_p95_ms",
        "clat_p50_ms",
    ]
    # Reorder columns if present
    cols = [c for c in preferred_order if c in df.columns] + [
        c for c in df.columns if c not in preferred_order
    ]
    df = df[cols]
    # Sort for readability
    sort_cols = [c for c in ("task", "runner", "timestamp", "bs_bytes", "iodepth", "source_file", "jobname", "op") if c in df.columns]
    if sort_cols:
        df = df.sort_values(by=sort_cols, kind="mergesort").reset_index(drop=True)
    return df


def _write_task_sheet(writer: "pd.ExcelWriter", task_name: str, df_task: "pd.DataFrame") -> None:
    sheet_names: Dict[str, int] = getattr(writer, "_sheet_name_registry", {})
    if not sheet_names:
        sheet_names = {}
        setattr(writer, "_sheet_name_registry", sheet_names)
    safe_name = sanitize_sheet_name(task_name, used=sheet_names)

    # Write task data
    df_task.to_excel(writer, index=False, sheet_name=safe_name)
    ws = writer.sheets[safe_name]

    # Summary by op inside the sheet for charting
    summary_cols = [c for c in ["iops", "bw_MBps"] if c in df_task.columns]
    if not summary_cols:
        return
    # Runner-level means: average across repeated attempts per runner/op, then sum across runners
    summary_runner = None
    if "runner" in df_task.columns and df_task["runner"].notna().any():
        # mean per runner/op
        runner_means = df_task.groupby(["runner", "op"], dropna=False)[summary_cols].mean().reset_index()
        # sum across runners per op
        summary_runner = runner_means.groupby("op", dropna=False)[summary_cols].sum().reset_index()

    # Raw sum across all rows per op
    summary_rowsum = df_task.groupby("op", dropna=False)[summary_cols].sum().reset_index()

    # Prefer runner-summed means to avoid double-counting repeats
    summary = summary_runner if summary_runner is not None else summary_rowsum

    start_row = len(df_task) + 2
    summary.to_excel(writer, index=False, sheet_name=safe_name, startrow=start_row, startcol=0)

    # Add simple bar charts for IOPS and BW if openpyxl chart is available
    if BarChart is None or Reference is None:
        return

    num_rows = len(summary)
    if num_rows == 0:
        return

    # Categories (op names) at column A of the summary block
    cat_ref = Reference(ws, min_col=1, min_row=start_row + 2, max_row=start_row + 1 + num_rows)

    # Chart for IOPS
    if "iops" in summary.columns:
        data_ref_iops = Reference(ws, min_col=2, min_row=start_row + 1, max_row=start_row + 1 + num_rows)
        chart_iops = BarChart()
        chart_iops.title = f"{task_name} - IOPS by op"
        chart_iops.y_axis.title = "IOPS (sum)"
        chart_iops.add_data(data_ref_iops, titles_from_data=True)
        chart_iops.set_categories(cat_ref)
        anchor_row = start_row + num_rows + 3
        ws.add_chart(chart_iops, ws.cell(row=anchor_row, column=1).coordinate)

    # Chart for BW
    if "bw_MBps" in summary.columns:
        data_ref_bw = Reference(ws, min_col=3 if "iops" in summary.columns else 2, min_row=start_row + 1, max_row=start_row + 1 + num_rows)
        chart_bw = BarChart()
        chart_bw.title = f"{task_name} - Bandwidth by op"
        chart_bw.y_axis.title = "MB/s (sum)"
        chart_bw.add_data(data_ref_bw, titles_from_data=True)
        chart_bw.set_categories(cat_ref)
        anchor_row = start_row + num_rows + 3
        ws.add_chart(chart_bw, ws.cell(row=anchor_row, column=10).coordinate)

    # Pivot by block size (mean metrics)
    start_row = start_row + num_rows + 16
    # Build runner-normalized means per bs/op: mean per runner first, then mean across runners
    metrics_bs = [c for c in ["iops", "bw_MBps", "clat_p99_ms", "clat_p95_ms", "clat_p50_ms"] if c in df_task.columns]
    if "runner" in df_task.columns and df_task["runner"].notna().any():
        tmp = df_task.groupby(["bs", "op", "runner"], dropna=False)[metrics_bs].mean().reset_index()
        pivot_bs = tmp.groupby(["bs", "op"], dropna=False)[metrics_bs].mean().reset_index()
    else:
        pivot_bs = df_task.groupby(["bs", "op"], dropna=False)[metrics_bs].mean().reset_index()
    pivot_bs.to_excel(writer, index=False, sheet_name=safe_name, startrow=start_row, startcol=0)

    # Bar chart: IOPS by block size grouped by op
    if LineChart is not None:
        pass  # no-op, we use BarChart for grouped bars below
    # Build a matrix bs x op for IOPS
    if "iops" in df_task.columns and len(pivot_bs) > 0:
        mat = pivot_bs.pivot_table(index="bs", columns="op", values="iops", aggfunc="mean").reset_index()
        mat_row = start_row + len(pivot_bs) + 3
        mat.to_excel(writer, index=False, sheet_name=safe_name, startrow=mat_row, startcol=0)
        # Build grouped bar chart
        try:
            num_rows_m = len(mat)
            num_cols_m = len(mat.columns)
            cat_ref2 = Reference(ws, min_col=1, min_row=mat_row + 2, max_row=mat_row + 1 + num_rows_m)
            data_ref2 = Reference(ws, min_col=2, max_col=num_cols_m, min_row=mat_row + 1, max_row=mat_row + 1 + num_rows_m)
            chart_bs = BarChart()
            chart_bs.title = f"{task_name} - IOPS by BS"
            chart_bs.y_axis.title = "IOPS (mean)"
            chart_bs.y_axis.number_format = "#,##0"
            chart_bs.add_data(data_ref2, titles_from_data=True)
            chart_bs.set_categories(cat_ref2)
            ws.add_chart(chart_bs, ws.cell(row=mat_row, column=10).coordinate)
        except Exception:
            pass

    # Pivot by iodepth (mean metrics) with line chart
    start_row2 = start_row + len(pivot_bs) + 20
    if "iodepth" in df_task.columns and df_task["iodepth"].notna().any():
        pivot_id = (
            df_task.groupby(["iodepth", "op"], dropna=False)[[c for c in ["iops", "bw_MBps"] if c in df_task.columns]]
            .mean()
            .reset_index()
        )
        pivot_id.to_excel(writer, index=False, sheet_name=safe_name, startrow=start_row2, startcol=0)
        # Line chart for IOPS over iodepth
        if LineChart is not None:
            try:
                mat2 = pivot_id.pivot_table(index="iodepth", columns="op", values="iops", aggfunc="mean").reset_index()
                mat2_row = start_row2 + len(pivot_id) + 3
                mat2.to_excel(writer, index=False, sheet_name=safe_name, startrow=mat2_row, startcol=0)
                num_rows_m2 = len(mat2)
                num_cols_m2 = len(mat2.columns)
                cat_ref3 = Reference(ws, min_col=1, min_row=mat2_row + 2, max_row=mat2_row + 1 + num_rows_m2)
                data_ref3 = Reference(ws, min_col=2, max_col=num_cols_m2, min_row=mat2_row + 1, max_row=mat2_row + 1 + num_rows_m2)
                chart_line = LineChart()
                chart_line.title = f"{task_name} - IOPS vs iodepth"
                chart_line.y_axis.title = "IOPS (mean)"
                chart_line.y_axis.number_format = "#,##0"
                chart_line.add_data(data_ref3, titles_from_data=True)
                chart_line.set_categories(cat_ref3)
                ws.add_chart(chart_line, ws.cell(row=mat2_row, column=10).coordinate)
            except Exception:
                pass

    # Heatmap: bs x iodepth for BW
    if ColorScaleRule is not None and {"bs", "iodepth", "bw_MBps"}.issubset(df_task.columns):
        try:
            heat = df_task.pivot_table(index="bs", columns="iodepth", values="bw_MBps", aggfunc="mean")
            heat_row = start_row2 + 40
            heat_col = 0
            heat.to_excel(writer, sheet_name=safe_name, startrow=heat_row, startcol=heat_col)
            # Apply 3-color scale to the data body
            min_row = heat_row + 2
            max_row = heat_row + 1 + len(heat.index)
            min_col = heat_col + 2
            max_col = heat_col + 1 + len(heat.columns)
            rng = f"{ws.cell(row=min_row, column=min_col).coordinate}:{ws.cell(row=max_row, column=max_col).coordinate}"
            rule = ColorScaleRule(start_type='min', start_color='FFF5F5F5', mid_type='percentile', mid_value=50, mid_color='FFBDD7EE', end_type='max', end_color='FF2E75B6')
            ws.conditional_formatting.add(rng, rule)
        except Exception:
            pass

    # Job spread stats
    if "jobname" in df_task.columns and df_task["jobname"].notna().any():
        try:
            spread = (
                df_task.groupby(["jobname", "op"], dropna=False)[[c for c in ["iops", "bw_MBps"] if c in df_task.columns]]
                .agg(['count', 'mean', 'min', 'max'])
            )
            # Flatten columns
            spread.columns = [f"{metric}_{stat}" for metric, stat in spread.columns.to_flat_index()]
            spread = spread.reset_index()
            spread_row = heat_row + 40
            spread.to_excel(writer, index=False, sheet_name=safe_name, startrow=spread_row, startcol=0)
        except Exception:
            pass


def export_to_excel(df: "pd.DataFrame", output_path: str, remove_suffix: bool = False, detailed_sheets: bool = False) -> None:
    if df.empty:
        with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
            empty = pd.DataFrame({"message": ["No results parsed"]})
            empty.to_excel(writer, index=False, sheet_name="summary")
        return

    # If requested, normalize task names across ALL outputs (not just summary sheet)
    if remove_suffix and "task" in df.columns:
        df = df.copy()
        df["task"] = df["task"].apply(
            lambda x: re.sub(r"-20\d{6}T\d{6}Z$", "", x) if isinstance(x, str) else x
        )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # Per-task sheets (detailed mode only)
        if detailed_sheets and "task" in df.columns:
            for task_name, df_task in df.groupby("task", dropna=False):
                df_task_sorted = df_task.reset_index(drop=True)
                _write_task_sheet(writer, str(task_name) if task_name is not None else "(no task)", df_task_sorted)
                print(f"[excel] task sheet: {task_name} rows={len(df_task_sorted)}")

        # Summary sheet: totals per task and per op, plus latency means
        summary_rows: List[Dict[str, object]] = []
        if {"task", "op"}.issubset(df.columns):
            # Optionally normalize task names by removing trailing -YYYYMMDDTHHMMSSZ
            df_summary = df.copy()
            tcol = "task"
            if remove_suffix and "task" in df_summary.columns:
                df_summary["task_norm"] = df_summary["task"].apply(
                    lambda x: re.sub(r"-20\d{6}T\d{6}Z$", "", x) if isinstance(x, str) else x
                )
                tcol = "task_norm"

            # Runner-normalized aggregation: mean per runner, then sum across runners per task/op (and access_pattern) for throughput; latency means averaged
            metrics_thr = [c for c in ["iops", "bw_MBps"] if c in df_summary.columns]
            metrics_lat = [c for c in ["clat_p99_ms", "clat_p95_ms", "clat_p50_ms"] if c in df_summary.columns]
            group_base = [tcol]
            if "access_pattern" in df_summary.columns:
                group_base.append("access_pattern")
            group_with_op = group_base + ["op"]
            if "runner" in df_summary.columns and df_summary["runner"].notna().any():
                tmp = df_summary.groupby(group_with_op + ["runner"], dropna=False)[metrics_thr + metrics_lat].mean().reset_index()
                # Sum throughput across runners, average latencies across runners
                agg_thr = tmp.groupby(group_with_op, dropna=False)[metrics_thr].sum()
                agg_lat = tmp.groupby(group_with_op, dropna=False)[metrics_lat].mean()
                agg = (
                    pd.concat([agg_thr, agg_lat], axis=1)
                    .reset_index()
                )
            else:
                agg = df_summary.groupby(group_with_op, dropna=False)[metrics_thr + metrics_lat].agg({**{m: "sum" for m in metrics_thr}, **{m: "mean" for m in metrics_lat}}).reset_index()
            # Build simple summary table with explicit columns:
            # iops_read, iops_write, iops_total_all, bw_MBps_read, bw_MBps_write, bw_total_all_MBps,
            # latency_p99_ms, latency_p95_ms, runs_count
            # If access_pattern is present, carry it into index for clarity
            index_cols = [c for c in [tcol, "access_pattern"] if c in agg.columns]
            iops_p = agg.pivot_table(index=index_cols, columns="op", values="iops", aggfunc="sum").fillna(0)
            bw_p = agg.pivot_table(index=index_cols, columns="op", values="bw_MBps", aggfunc="sum").fillna(0)
            # Optional latency (p99/p95) as average across ops
            lat_p99 = agg.pivot_table(index=tcol, columns="op", values="clat_p99_ms", aggfunc="mean") if "clat_p99_ms" in agg.columns else None
            lat_p95 = agg.pivot_table(index=tcol, columns="op", values="clat_p95_ms", aggfunc="mean") if "clat_p95_ms" in agg.columns else None

            # Normalize column names and combine
            def _rename(prefix_df, prefix):
                if prefix_df is None:
                    return None
                renamed = prefix_df.copy()
                renamed.columns = [f"{prefix}_{str(c)}" for c in renamed.columns]
                return renamed

            iops_df = _rename(iops_p, "iops")
            bw_df = _rename(bw_p, "bw_MBps")

            summary_df = None
            if iops_df is not None and bw_df is not None:
                summary_df = iops_df.join(bw_df, how="outer").fillna(0)
            elif iops_df is not None:
                summary_df = iops_df.copy()
            elif bw_df is not None:
                summary_df = bw_df.copy()
            else:
                summary_df = pd.DataFrame(index=agg[tcol].drop_duplicates())

            # Totals
            for total_col, prefix in (("iops_total_all", "iops_"), ("bw_total_all_MBps", "bw_MBps_")):
                cols = [c for c in summary_df.columns if c.startswith(prefix)]
                summary_df[total_col] = summary_df[cols].sum(axis=1) if cols else 0

            # Latency p99/p95 combined (mean across ops)
            if lat_p99 is not None:
                summary_df["latency_p99_ms"] = lat_p99.mean(axis=1)
            else:
                summary_df["latency_p99_ms"] = float("nan")
            if lat_p95 is not None:
                summary_df["latency_p95_ms"] = lat_p95.mean(axis=1)
            else:
                summary_df["latency_p95_ms"] = float("nan")

            # Overall mean latency (weighted by total_ios), and per-runner mean latency
            latency_mean_ms = []
            latency_mean_ms_per_runner = []
            if {"lat_mean_ms", "total_ios"}.issubset(df_summary.columns):
                # Runner-level aggregation
                tmp_groups = [tcol]
                if "access_pattern" in df_summary.columns:
                    tmp_groups.append("access_pattern")
                tmp_lat = df_summary.groupby(tmp_groups + ["op", "runner"], dropna=False)[["lat_mean_ms", "total_ios"]].agg({"lat_mean_ms": "mean", "total_ios": "sum"}).reset_index()

                # Weighted across ops and runners for each task grouping
                for task_name, g in tmp_lat.groupby(tmp_groups, dropna=False):
                    w = g["total_ios"].fillna(0)
                    vals = g["lat_mean_ms"].astype(float)
                    total_w = w.sum()
                    if total_w and total_w > 0:
                        latency_mean_ms.append((task_name, float((vals * w).sum() / total_w)))
                    else:
                        latency_mean_ms.append((task_name, float("nan")))

                # Per-runner mean: first weight by ios across ops within runner, then average runners
                def _runner_weighted_mean(grp):
                    w = grp["total_ios"].fillna(0)
                    vals = grp["lat_mean_ms"].astype(float)
                    tw = w.sum()
                    return float((vals * w).sum() / tw) if tw and tw > 0 else float("nan")

                runner_lat = (
                    tmp_lat.groupby(tmp_groups + ["runner"], dropna=False)[["lat_mean_ms", "total_ios"]]
                    .apply(_runner_weighted_mean)
                    .reset_index(name="lat_mean_ms_runner")
                )
                pr = runner_lat.groupby(tmp_groups, dropna=False)["lat_mean_ms_runner"].mean().reset_index()

                # Merge back to summary_df index
                key_cols = tmp_groups
                lm_cols = key_cols + ["latency_mean_ms"]
                # Expand keys into columns even when keys are tuples
                lm_rows = []
                for keys, val in latency_mean_ms:
                    if not isinstance(keys, tuple):
                        keys = (keys,)
                    lm_rows.append(list(keys) + [val])
                lm_df = pd.DataFrame(lm_rows, columns=lm_cols)
                lm_df = lm_df.merge(pr.rename(columns={"lat_mean_ms_runner": "latency_mean_ms_per_runner"}), on=key_cols, how="left")
                # Merge on the same index columns used by summary_df
                summary_df = summary_df.merge(lm_df.set_index(index_cols), left_index=True, right_index=True, how="left")
            else:
                summary_df["latency_mean_ms"] = float("nan")
                summary_df["latency_mean_ms_per_runner"] = float("nan")

            # Runs count (distinct runners per task if available; else row count)
            if "runner" in df_summary.columns and df_summary["runner"].notna().any():
                rc = df_summary.groupby(tcol)["runner"].nunique().reset_index(name="runs_count")
            else:
                rc = df_summary.groupby(tcol).size().reset_index(name="runs_count")
            rc = rc.rename(columns={tcol: "task"})

            # Ensure presence of expected columns even if missing
            for col in [
                "iops_read", "iops_write", "bw_MBps_read", "bw_MBps_write",
            ]:
                if col not in summary_df.columns:
                    summary_df[col] = 0

            # Ensure task present; keep access_pattern if available
            summary_df = summary_df.reset_index().rename(columns={tcol: "task"})
            summary_df = summary_df.merge(rc, on="task", how="left")

            # Per-runner averages (throughput: divide sums by runs_count; latency already averaged per runner)
            def _div(a, b):
                try:
                    return a / b if b and b != 0 else float("nan")
                except Exception:
                    return float("nan")

            for base_col in ["iops_read", "iops_write", "bw_MBps_read", "bw_MBps_write"]:
                per_col = f"{base_col}_per_runner"
                summary_df[per_col] = summary_df.apply(lambda r: _div(r.get(base_col, float("nan")), r.get("runs_count", 0)), axis=1)

            # Order columns as requested
            ordered = [
                "task",
                "access_pattern" if "access_pattern" in summary_df.columns else None,
                "iops_read",
                "iops_write",
                "iops_total_all",
                "iops_read_per_runner",
                "iops_write_per_runner",
                "bw_MBps_read",
                "bw_MBps_write",
                "bw_total_all_MBps",
                "bw_MBps_read_per_runner",
                "bw_MBps_write_per_runner",
                "latency_mean_ms",
                "latency_mean_ms_per_runner",
                "latency_p99_ms",
                "latency_p95_ms",
                "runs_count",
            ]
            ordered = [c for c in ordered if c is not None]

            # Sort tasks in desired order and write
            summary_sorted = _sort_df_by_task_order(summary_df, task_col="task")
            summary_sorted[ordered].to_excel(writer, index=False, sheet_name="summary")
            print(f"[excel] summary rows={len(summary_sorted)} cols={len(ordered)}")
            ws_sum = writer.sheets["summary"]
            # Add run counts per task
            try:
                counts = df_summary.groupby(tcol).size().reset_index(name="row_count").rename(columns={tcol: "task"})
                counts = _sort_df_by_task_order(counts, task_col="task")
                counts.to_excel(writer, index=False, sheet_name="summary", startrow=len(summary_sorted) + 3, startcol=0)
            except Exception:
                pass

            # Charts for summary: grouped bars for IOPS_* and BW_* columns
            if BarChart is not None and Reference is not None:
                try:
                    col_names = ordered
                    n_rows = len(summary_sorted) + 1  # including header
                    # Categories are tasks in column 1 (A), rows 2..n_rows
                    cat_ref = Reference(ws_sum, min_col=1, min_row=2, max_row=n_rows)

                    # IOPS columns (contiguous by our ordering)
                    iops_first = next((i for i, c in enumerate(col_names, start=1) if c.startswith("iops_")), None)
                    iops_last = max([i for i, c in enumerate(col_names, start=1) if c.startswith("iops_")], default=None)
                    if iops_first is not None and iops_last is not None and iops_last >= iops_first:
                        data_ref_iops = Reference(ws_sum, min_col=iops_first, max_col=iops_last, min_row=1, max_row=n_rows)
                        chart_iops = BarChart()
                        chart_iops.title = "Total IOPS per task (by op)"
                        chart_iops.y_axis.title = "IOPS (sum)"
                        chart_iops.y_axis.number_format = "#,##0"
                        chart_iops.add_data(data_ref_iops, titles_from_data=True)
                        chart_iops.set_categories(cat_ref)
                        ws_sum.add_chart(chart_iops, "J2")

                    # BW columns (contiguous by our ordering)
                    bw_first = next((i for i, c in enumerate(col_names, start=1) if c.startswith("bw_MBps_")), None)
                    bw_last = max([i for i, c in enumerate(col_names, start=1) if c.startswith("bw_MBps_")], default=None)
                    if bw_first is not None and bw_last is not None and bw_last >= bw_first:
                        data_ref_bw = Reference(ws_sum, min_col=bw_first, max_col=bw_last, min_row=1, max_row=n_rows)
                        chart_bw = BarChart()
                        chart_bw.title = "Total Bandwidth per task (by op)"
                        chart_bw.y_axis.title = "MB/s (sum)"
                        chart_bw.y_axis.number_format = "0.0"
                        chart_bw.add_data(data_ref_bw, titles_from_data=True)
                        chart_bw.set_categories(cat_ref)
                        ws_sum.add_chart(chart_bw, "J20")
                except Exception:
                    pass

            # Summary detailed: speed per block size (runner-normalized)
            try:
                # Build runner-normalized aggregation per task×bs×op
                metrics_thr = [c for c in ["iops", "bw_MBps"] if c in df_summary.columns]
                metrics_lat = [c for c in ["clat_p99_ms", "clat_p95_ms"] if c in df_summary.columns]
                base_groups = [tcol]
                if "access_pattern" in df_summary.columns:
                    base_groups.append("access_pattern")
                group_keys_with_op = base_groups + ["bs", "bs_bytes", "op"]
                if "runner" in df_summary.columns and df_summary["runner"].notna().any():
                    tmp = df_summary.groupby(group_keys_with_op + ["runner"], dropna=False)[metrics_thr + metrics_lat].mean().reset_index()
                    agg_d_thr = tmp.groupby(group_keys_with_op, dropna=False)[metrics_thr].sum()
                    agg_d_lat = tmp.groupby(group_keys_with_op, dropna=False)[metrics_lat].mean()
                    agg_d = (
                        pd.concat([agg_d_thr, agg_d_lat], axis=1)
                        .reset_index()
                    )
                    rc_groups = base_groups + ["bs"]
                    runs_cnt = df_summary.groupby(rc_groups)['runner'].nunique().reset_index(name='runs_count')
                else:
                    agg_d = df_summary.groupby(group_keys_with_op, dropna=False)[metrics_thr + metrics_lat].agg({**{m: "sum" for m in metrics_thr}, **{m: "mean" for m in metrics_lat}}).reset_index()
                    rc_groups = base_groups + ["bs"]
                    runs_cnt = df_summary.groupby(rc_groups).size().reset_index(name='runs_count')

                # Build simple per-bs table
                index_keys = base_groups + ["bs", "bs_bytes"]
                iops_bs = agg_d.pivot_table(index=index_keys, columns="op", values="iops", aggfunc="sum").fillna(0)
                bw_bs = agg_d.pivot_table(index=index_keys, columns="op", values="bw_MBps", aggfunc="sum").fillna(0)
                lat99_bs = agg_d.pivot_table(index=index_keys, columns="op", values="clat_p99_ms", aggfunc="mean") if "clat_p99_ms" in agg_d.columns else None
                lat95_bs = agg_d.pivot_table(index=index_keys, columns="op", values="clat_p95_ms", aggfunc="mean") if "clat_p95_ms" in agg_d.columns else None

                def _rn(df_in, prefix):
                    if df_in is None:
                        return None
                    out = df_in.copy()
                    out.columns = [f"{prefix}_{str(c)}" for c in out.columns]
                    return out

                tbl = None
                parts = [_rn(iops_bs, "iops"), _rn(bw_bs, "bw_MBps")]
                base = None
                for p in parts:
                    if p is not None:
                        base = p if base is None else base.join(p, how="outer")
                if base is None:
                    base = pd.DataFrame(index=iops_bs.index if isinstance(iops_bs, pd.DataFrame) else [])

                # totals
                for total_col, prefix in (("iops_total_all", "iops_"), ("bw_total_all_MBps", "bw_MBps_")):
                    cols = [c for c in base.columns if c.startswith(prefix)]
                    base[total_col] = base[cols].sum(axis=1) if cols else 0

                # Latency aggregates across ops
                if lat99_bs is not None:
                    base["latency_p99_ms"] = lat99_bs.mean(axis=1)
                else:
                    base["latency_p99_ms"] = float("nan")
                if lat95_bs is not None:
                    base["latency_p95_ms"] = lat95_bs.mean(axis=1)
                else:
                    base["latency_p95_ms"] = float("nan")

                base = base.reset_index().rename(columns={tcol: "task"})
                # Attach runs count
                runs_cnt = runs_cnt.rename(columns={tcol: "task"})
                merge_keys = ["task", "bs"]
                if "access_pattern" in runs_cnt.columns and "access_pattern" in base.columns:
                    merge_keys = ["task", "access_pattern", "bs"]
                base = base.merge(runs_cnt, on=merge_keys, how="left")

                # Per-runner averages for detailed (throughput divided by runs_count)
                def _div(a, b):
                    try:
                        return a / b if b and b != 0 else float("nan")
                    except Exception:
                        return float("nan")
                for base_col in ["iops_read", "iops_write", "bw_MBps_read", "bw_MBps_write"]:
                    per_col = f"{base_col}_per_runner"
                    if base_col in base.columns:
                        base[per_col] = base.apply(lambda r: _div(r.get(base_col, float("nan")), r.get("runs_count", 0)), axis=1)


                # Ensure required columns exist
                for col in ["iops_read", "iops_write", "bw_MBps_read", "bw_MBps_write"]:
                    if col not in base.columns:
                        base[col] = 0

                # Sort tasks by degree 3,6,9,12,15 and bs ascending
                desired = [3, 6, 9, 12, 15]
                base["__deg__"] = base["task"].map(_extract_parallel_degree)
                base["__rank__"] = base["__deg__"].map(lambda x: desired.index(x) if x in desired else len(desired))
                base = base.sort_values(by=["__rank__", "__deg__", "task", "bs_bytes"], kind="mergesort").drop(columns=["__rank__"]) 

                # Clarify which ops are present in this row (read/write)
                def _ops_present(row):
                    flags = []
                    try:
                        if float(row.get("iops_read", 0) or 0) > 0 or float(row.get("bw_MBps_read", 0) or 0) > 0:
                            flags.append("read")
                        if float(row.get("iops_write", 0) or 0) > 0 or float(row.get("bw_MBps_write", 0) or 0) > 0:
                            flags.append("write")
                    except Exception:
                        pass
                    return ",".join(flags)
                base["ops"] = base.apply(_ops_present, axis=1)

                # Add a human-friendly label: "<runners> runners - <bs>, <ops>"
                try:
                    base["runners"] = base["task"].map(_extract_parallel_degree)
                except Exception:
                    base["runners"] = None
                def _make_label(row):
                    deg = row.get("runners")
                    bs_val = row.get("bs")
                    ops_val = row.get("ops") or ""
                    try:
                        deg_str = f"{int(deg)}" if deg is not None else ""
                    except Exception:
                        deg_str = str(deg) if deg is not None else ""
                    prefix = f"{deg_str} runners" if deg_str else "runners"
                    return f"{prefix} - {bs_val}, {ops_val}".strip()
                base["label"] = base.apply(_make_label, axis=1)
                # Also provide a split-friendly variant without ops
                def _make_runners_bs(row):
                    deg = row.get("runners")
                    bs_val = row.get("bs")
                    try:
                        deg_str = f"{int(deg)}" if deg is not None else ""
                    except Exception:
                        deg_str = str(deg) if deg is not None else ""
                    prefix = f"{deg_str} runners" if deg_str else "runners"
                    return f"{prefix} - {bs_val}".strip()
                base["runners_bs"] = base.apply(_make_runners_bs, axis=1)

                # Final column order
                detailed_cols = [
                    "runners_bs", "ops",
                    "task", "access_pattern" if "access_pattern" in base.columns else None, "bs",
                    "iops_read", "iops_write", "iops_total_all",
                    "iops_read_per_runner", "iops_write_per_runner",
                    "bw_MBps_read", "bw_MBps_write", "bw_total_all_MBps",
                    "bw_MBps_read_per_runner", "bw_MBps_write_per_runner",
                    "latency_p99_ms", "latency_p95_ms",
                    "runs_count"
                ]
                ordered_cols = [c for c in detailed_cols if c is not None]
                base[ordered_cols].to_excel(writer, index=False, sheet_name="summary_detailed")
                print(f"[excel] summary_detailed rows={len(base)} cols={len(ordered_cols)}")
                # Also export summary_detailed to CSV for offline inspection
                try:
                    csv_path = os.path.splitext(output_path)[0] + "_summary_detailed.csv"
                    base[ordered_cols].to_csv(csv_path, index=False)
                    print(f"[csv] wrote {csv_path} rows={len(base)} cols={len(ordered_cols)}")
                except Exception as e:
                    print(f"[csv] error writing summary_detailed CSV: {e}")

                # Also create per-block-size sheets mirroring summary_detailed, one sheet per bs
                bs_values = sorted([str(b) for b in base["bs"].dropna().unique().tolist()]) if "bs" in base.columns else []
                for bs_value in bs_values:
                    subset = base[base["bs"].astype(str) == bs_value].reset_index(drop=True)
                    # Ensure unique and valid sheet name
                    sheet_names: Dict[str, int] = getattr(writer, "_sheet_name_registry", {})
                    if not sheet_names:
                        sheet_names = {}
                        setattr(writer, "_sheet_name_registry", sheet_names)
                    sheet_label = sanitize_sheet_name(f"bs_{bs_value}", used=sheet_names)
                    subset[ordered_cols].to_excel(writer, index=False, sheet_name=sheet_label)
                    print(f"[excel] bs sheet: {sheet_label} rows={len(subset)} cols={len(ordered_cols)}")
            except Exception:
                pass
        else:
            # Fallback: write full df if task/op missing
            df.to_excel(writer, index=False, sheet_name="summary")

        # New: summary_per_runner sheet for explicit task×runner split (detailed only)
        if detailed_sheets and {"task", "runner", "op"}.issubset(df.columns) and df["runner"].notna().any():
            try:
                metrics_thr = [c for c in ["iops", "bw_MBps"] if c in df.columns]
                metrics_lat = [c for c in ["clat_p99_ms", "clat_p95_ms", "clat_p50_ms"] if c in df.columns]
                per_runner = df.groupby(["task", "runner", "op"], dropna=False)[metrics_thr + metrics_lat].mean().reset_index()
                per_runner_pivot = per_runner.pivot_table(index=["task", "runner"], columns="op", values=metrics_thr + metrics_lat)
                # Sort by task order
                pr_sorted = _sort_df_by_task_order(per_runner_pivot.reset_index(), task_col="task").set_index(["task", "runner"])
                pr_sorted.to_excel(writer, sheet_name="summary_per_runner")
                print(f"[excel] summary_per_runner rows={len(pr_sorted)}")
            except Exception:
                pass

        # Cross-task block-size comparison sheet (detailed only)
        if detailed_sheets and {"task", "bs", "op"}.issubset(df.columns):
            try:
                # Include access_pattern if present to separate sequential vs random contexts
                metrics_cols = [c for c in ["iops", "bw_MBps"] if c in df.columns]
                group_cols = ["task", "bs", "op"] + (["access_pattern"] if "access_pattern" in df.columns else [])
                comp = df.groupby(group_cols, dropna=False)[metrics_cols].mean().reset_index()
                index_cols = ["task"] + (["access_pattern"] if "access_pattern" in comp.columns else [])
                comp_pivot = comp.pivot_table(index=index_cols, columns=["bs", "op"], values=metrics_cols).sort_index(axis=1)
                # Sort task index using the desired order
                comp_pivot_sorted = comp_pivot.copy()
                comp_pivot_sorted = _sort_df_by_task_order(comp_pivot_sorted.reset_index(), task_col="task")
                comp_pivot_sorted = comp_pivot_sorted.set_index(index_cols)
                comp_pivot_sorted.to_excel(writer, sheet_name="bs_compare")
                print(f"[excel] bs_compare rows={len(comp_pivot_sorted)}")

                # Flattened version for easier charting
                comp_flat = comp_pivot_sorted.copy()
                comp_flat.columns = [f"{metric}_{bs}_{op}" for metric, bs, op in comp_pivot.columns.to_flat_index()]
                comp_flat = comp_flat.reset_index()
                # Reorder columns: task, access_pattern (if present), then metrics
                cols_iops = [c for c in comp_flat.columns if c.startswith("iops_")]
                cols_bw = [c for c in comp_flat.columns if c.startswith("bw_")]
                leading_cols = [c for c in ["task", "access_pattern"] if c in comp_flat.columns]
                comp_flat = comp_flat[leading_cols + sorted(cols_iops) + sorted(cols_bw)]
                ws_bc = writer.sheets["bs_compare"]
                flat_row = comp_pivot.shape[0] + 3
                comp_flat.to_excel(writer, sheet_name="bs_compare", startrow=flat_row, index=False)
                print(f"[excel] bs_compare flattened rows={len(comp_flat)} at row={flat_row}")

                # Charts: grouped bars for IOPS across bs/op columns
                if BarChart is not None and Reference is not None and len(cols_iops) > 0:
                    try:
                        n_rows = comp_flat.shape[0] + 1
                        n_cols = 1 + len(cols_iops)  # task + iops cols
                        cat_ref = Reference(ws_bc, min_col=1, min_row=flat_row + 2, max_row=flat_row + n_rows)
                        data_ref = Reference(ws_bc, min_col=2, max_col=1 + len(cols_iops), min_row=flat_row + 1, max_row=flat_row + n_rows)
                        chart = BarChart()
                        chart.title = "IOPS by task (grouped by bs/op)"
                        chart.y_axis.title = "IOPS (mean)"
                        chart.y_axis.number_format = "#,##0"
                        chart.add_data(data_ref, titles_from_data=True)
                        chart.set_categories(cat_ref)
                        ws_bc.add_chart(chart, f"J2")
                    except Exception:
                        pass

                # Chart for BW
                if BarChart is not None and Reference is not None and len(cols_bw) > 0:
                    try:
                        # BW table starts at column 2 + len(iops)
                        bw_start_col = 2 + len(cols_iops)
                        bw_end_col = bw_start_col + len(cols_bw) - 1
                        n_rows = comp_flat.shape[0] + 1
                        cat_ref = Reference(ws_bc, min_col=1, min_row=flat_row + 2, max_row=flat_row + n_rows)
                        data_ref = Reference(ws_bc, min_col=bw_start_col, max_col=bw_end_col, min_row=flat_row + 1, max_row=flat_row + n_rows)
                        chart_bw = BarChart()
                        chart_bw.title = "BW by task (grouped by bs/op)"
                        chart_bw.y_axis.title = "MB/s (mean)"
                        chart_bw.y_axis.number_format = "0.0"
                        chart_bw.add_data(data_ref, titles_from_data=True)
                        chart_bw.set_categories(cat_ref)
                        ws_bc.add_chart(chart_bw, f"J20")
                    except Exception:
                        pass
            except Exception:
                pass

        # Per-block-size summary tabs: one sheet per block size value (legacy generator disabled; bs_* are created from summary_detailed above)
        if detailed_sheets and False and {"bs", "op"}.issubset(df.columns):
            try:
                for bs_value, df_bs in df.groupby("bs", dropna=False):
                    sheet_label = f"bs_{str(bs_value)}"
                    # Build a compact summary per task (and access_pattern if present)
                    # Put access_pattern into the index so it appears as a dedicated column
                    idx_cols = [c for c in ["task", "access_pattern"] if c in df_bs.columns]
                    col_group = ["op"]
                    thr_cols = [c for c in ["iops", "bw_MBps"] if c in df_bs.columns]
                    lat_cols = [c for c in ["clat_p99_ms", "clat_p95_ms"] if c in df_bs.columns]
                    # Runner-normalized: mean per runner, then sum across runners for throughput and mean for latency
                    if "runner" in df_bs.columns and df_bs["runner"].notna().any():
                        tmp = df_bs.groupby(idx_cols + col_group + ["runner"], dropna=False)[thr_cols + lat_cols].mean().reset_index()
                        agg_thr = tmp.groupby(idx_cols + col_group, dropna=False)[thr_cols].sum()
                        agg_lat = tmp.groupby(idx_cols + col_group, dropna=False)[lat_cols].mean()
                        agg_bs = pd.concat([agg_thr, agg_lat], axis=1).reset_index()
                    else:
                        agg_bs = df_bs.groupby(idx_cols + col_group, dropna=False)[thr_cols + lat_cols].agg({**{m: "sum" for m in thr_cols}, **{m: "mean" for m in lat_cols}}).reset_index()

                    # Flatten to columns similar to summary
                    iops_p = agg_bs.pivot_table(index=idx_cols, columns=col_group, values="iops", aggfunc="sum").fillna(0)
                    bw_p = agg_bs.pivot_table(index=idx_cols, columns=col_group, values="bw_MBps", aggfunc="sum").fillna(0)
                    lat99_p = agg_bs.pivot_table(index=idx_cols, columns=col_group, values="clat_p99_ms", aggfunc="mean") if "clat_p99_ms" in agg_bs.columns else None
                    lat95_p = agg_bs.pivot_table(index=idx_cols, columns=col_group, values="clat_p95_ms", aggfunc="mean") if "clat_p95_ms" in agg_bs.columns else None

                    def _rn(df_in, prefix):
                        if df_in is None:
                            return None
                        out = df_in.copy()
                        # Flatten column labels; for op-only columns we just append the op name
                        new_cols = []
                        for c in out.columns:
                            if isinstance(c, tuple):
                                new_cols.append(f"{prefix}_" + "_".join(str(x) for x in c))
                            else:
                                new_cols.append(f"{prefix}_{str(c)}")
                        out.columns = new_cols
                        return out

                    table = None
                    parts = [_rn(iops_p, "iops"), _rn(bw_p, "bw_MBps")]
                    for p in parts:
                        if p is not None:
                            table = p if table is None else table.join(p, how="outer")
                    if table is None:
                        table = pd.DataFrame(index=iops_p.index if isinstance(iops_p, pd.DataFrame) else [])
                    # Totals
                    for total_col, prefix in (("iops_total_all", "iops_"), ("bw_total_all_MBps", "bw_MBps_")):
                        cols = [c for c in table.columns if c.startswith(prefix)]
                        table[total_col] = table[cols].sum(axis=1) if cols else 0
                    # Latency means
                    if lat99_p is not None:
                        table["latency_p99_ms"] = lat99_p.mean(axis=1)
                    if lat95_p is not None:
                        table["latency_p95_ms"] = lat95_p.mean(axis=1)

                    # Finalize
                    table = table.reset_index()
                    # Sort by task order
                    if "task" in table.columns:
                        table = _sort_df_by_task_order(table, task_col="task")
                    # Column order
                    leading = [c for c in ["task", "access_pattern"] if c in table.columns]
                    iops_cols = sorted([c for c in table.columns if c.startswith("iops_")])
                    bw_cols = sorted([c for c in table.columns if c.startswith("bw_MBps_")])
                    tail = [c for c in ["iops_total_all", "bw_total_all_MBps", "latency_p99_ms", "latency_p95_ms"] if c in table.columns]
                    ordered = leading + iops_cols + (["iops_total_all"] if "iops_total_all" in table.columns else []) + bw_cols + (["bw_total_all_MBps"] if "bw_total_all_MBps" in table.columns else []) + [c for c in ["latency_p99_ms", "latency_p95_ms"] if c in table.columns]
                    # Remove duplicates that list multiplication might have introduced when False
                    seen = set()
                    ordered = [x for x in ordered if not (x in seen or seen.add(x))]
                    table[ordered].to_excel(writer, index=False, sheet_name=sheet_label)
                    print(f"[excel] bs sheet: {sheet_label} rows={len(table)} cols={len(ordered)}")
            except Exception as e:
                print(f"[excel] bs sheets: error {e}")

        # Latency percentiles focus sheet (detailed only)
        if detailed_sheets and {"task", "op"}.issubset(df.columns) and {"clat_p99_ms", "clat_p95_ms", "clat_p50_ms"}.intersection(df.columns):
            try:
                lat = df.groupby(["task", "op"], dropna=False)[[c for c in ["clat_p99_ms", "clat_p95_ms", "clat_p50_ms"] if c in df.columns]].mean().reset_index()
                lat_pivot = lat.pivot_table(index="task", columns="op", values=[c for c in ["clat_p99_ms", "clat_p95_ms", "clat_p50_ms"] if c in df.columns])
                # Sort task order
                lat_pivot_sorted = _sort_df_by_task_order(lat_pivot.reset_index(), task_col="task").set_index("task")
                lat_pivot_sorted.to_excel(writer, sheet_name="latency")

                # Flattened version for charts
                lat_flat = lat_pivot_sorted.copy()
                lat_flat.columns = [f"{metric}_{op}" for metric, op in lat_pivot.columns.to_flat_index()]
                lat_flat = lat_flat.reset_index()
                # Order: task, p99_*, p95_*, p50_*
                cols_p99 = [c for c in lat_flat.columns if c.startswith("clat_p99_ms_")]
                cols_p95 = [c for c in lat_flat.columns if c.startswith("clat_p95_ms_")]
                cols_p50 = [c for c in lat_flat.columns if c.startswith("clat_p50_ms_")]
                lat_flat = lat_flat[["task"] + sorted(cols_p99) + sorted(cols_p95) + sorted(cols_p50)]
                ws_lat = writer.sheets["latency"]
                flat_row = lat_pivot.shape[0] + 3
                lat_flat.to_excel(writer, sheet_name="latency", startrow=flat_row, index=False)

                # Charts: bars for p99 and p95
                if BarChart is not None and Reference is not None and len(cols_p99) > 0:
                    try:
                        n_rows = lat_flat.shape[0] + 1
                        # p99 at columns 2..(1+len(cols_p99))
                        p99_first = 2
                        p99_last = 1 + len(cols_p99)
                        cat_ref = Reference(ws_lat, min_col=1, min_row=flat_row + 2, max_row=flat_row + n_rows)
                        data_ref = Reference(ws_lat, min_col=p99_first, max_col=p99_last, min_row=flat_row + 1, max_row=flat_row + n_rows)
                        chart_p99 = BarChart()
                        chart_p99.title = "Latency p99 (ms) by task"
                        chart_p99.y_axis.title = "ms"
                        chart_p99.y_axis.number_format = "0.00"
                        chart_p99.add_data(data_ref, titles_from_data=True)
                        chart_p99.set_categories(cat_ref)
                        ws_lat.add_chart(chart_p99, "J2")
                    except Exception:
                        pass

                if BarChart is not None and Reference is not None and len(cols_p95) > 0:
                    try:
                        # p95 columns follow p99 in our ordering
                        p95_first = 2 + len(cols_p99)
                        p95_last = p95_first + len(cols_p95) - 1
                        n_rows = lat_flat.shape[0] + 1
                        cat_ref = Reference(ws_lat, min_col=1, min_row=flat_row + 2, max_row=flat_row + n_rows)
                        data_ref = Reference(ws_lat, min_col=p95_first, max_col=p95_last, min_row=flat_row + 1, max_row=flat_row + n_rows)
                        chart_p95 = BarChart()
                        chart_p95.title = "Latency p95 (ms) by task"
                        chart_p95.y_axis.title = "ms"
                        chart_p95.y_axis.number_format = "0.00"
                        chart_p95.add_data(data_ref, titles_from_data=True)
                        chart_p95.set_categories(cat_ref)
                        ws_lat.add_chart(chart_p95, "J20")
                    except Exception:
                        pass
            except Exception:
                pass

        # Stability/variance sheet (detailed only)
        if detailed_sheets and {"task", "op", "bs", "iodepth"}.issubset(df.columns):
            try:
                metrics = [c for c in ["iops", "bw_MBps", "clat_p99_ms"] if c in df.columns]
                if metrics:
                    grp_cols = ["task", "op", "bs", "iodepth"]
                    stab = df.groupby(grp_cols)[metrics].agg(['count', 'mean', 'std', 'min', 'max'])
                    # Coefficient of variation
                    for m in metrics:
                        if (m, 'mean') in stab.columns and (m, 'std') in stab.columns:
                            stab[(m, 'cv_pct')] = (stab[(m, 'std')] / stab[(m, 'mean')]).replace([float('inf')], float('nan')) * 100.0
                    stab = stab.reset_index()
                    # Flatten
                    stab.columns = [
                        "_".join([c for c in col if c]) if isinstance(col, tuple) else col for col in stab.columns.to_flat_index()
                    ]
                    stab.to_excel(writer, index=False, sheet_name="stability")
            except Exception:
                pass

        # Timeline sheet (if timestamps present) (detailed only)
        if detailed_sheets and "timestamp" in df.columns and df["timestamp"].notna().any():
            try:
                tl = df.copy()
                tl = tl.sort_values("timestamp")
                # Aggregate per timestamp and op (sum IOPS/BW)
                tl_agg = tl.groupby(["timestamp", "op"], dropna=False)[[c for c in ["iops", "bw_MBps"] if c in tl.columns]].sum().reset_index()
                tl_pivot = tl_agg.pivot_table(index="timestamp", columns="op", values=[c for c in ["iops", "bw_MBps"] if c in tl.columns])
                tl_pivot.to_excel(writer, sheet_name="timeline")
                if LineChart is not None and Reference is not None:
                    ws_tl = writer.sheets["timeline"]
                    # Build a line chart for IOPS over time if available
                    # Assume headers at row 1 and timestamp in col A
                    max_row = ws_tl.max_row
                    max_col = ws_tl.max_column
                    if max_row > 2 and max_col > 2:
                        cat_ref = Reference(ws_tl, min_col=1, min_row=2, max_row=max_row)
                        data_ref = Reference(ws_tl, min_col=2, min_row=1, max_col=max_col, max_row=max_row)
                        chart_tl = LineChart()
                        chart_tl.title = "IOPS/BW over time"
                        chart_tl.add_data(data_ref, titles_from_data=True)
                        chart_tl.set_categories(cat_ref)
                        ws_tl.add_chart(chart_tl, "J2")
            except Exception:
                pass


def main() -> int:
    parser = argparse.ArgumentParser(description="Parse fio outputs and export to Excel")
    parser.add_argument(
        "--input",
        default=os.getcwd(),
        help="Directory containing fio output files (default: current directory)",
    )
    parser.add_argument(
        "--include",
        default=",".join(DEFAULT_EXTENSIONS),
        help="Comma-separated list of file extensions to include (default: json,txt,log,out)",
    )
    parser.add_argument(
        "--recurse",
        action="store_true",
        help="Recurse into subdirectories",
    )
    parser.add_argument(
        "--output",
        default="fio_summary.xlsx",
        help="Output Excel file path (default: fio_summary.xlsx)",
    )
    parser.add_argument(
        "--remove-suffix",
        action="store_true",
        help="On the summary sheet, remove trailing -YYYYMMDDTHHMMSSZ from task names",
    )
    parser.add_argument(
        "--detailed",
        action="store_true",
        help="Include detailed sheets (per-task, per-runner, bs_compare, latency, stability, timeline)",
    )

    args = parser.parse_args()

    input_dir = os.path.abspath(args.input)
    if not os.path.isdir(input_dir):
        sys.stderr.write(f"Input is not a directory: {input_dir}\n")
        return 2

    exts = [e.strip().lstrip(".") for e in args.include.split(",") if e.strip()]
    files = list(iter_files(input_dir, exts, args.recurse))
    if not files:
        sys.stderr.write("No files found matching the given extensions.\n")
        # Still create an empty excel to signal completion
        export_to_excel(pd.DataFrame(), args.output)
        print(f"Wrote empty Excel to {args.output}")
        return 0

    all_rows: List[ParsedRow] = []
    for path in files:
        all_rows.extend(parse_file(path, input_root=input_dir))

    df = rows_to_dataframe(all_rows)
    export_to_excel(df, args.output, remove_suffix=args.remove_suffix, detailed_sheets=args.detailed)
    print(f"Parsed {len(files)} file(s), produced {len(df)} row(s): {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())


